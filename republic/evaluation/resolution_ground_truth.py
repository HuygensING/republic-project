from typing import Dict, List, Generator, Set
import os
import glob

from openpyxl import Workbook, load_workbook
from elasticsearch import Elasticsearch

import republic.elastic.republic_elasticsearch as rep_es
from republic.model.republic_document_model import ResolutionPageDoc, Resolution


#########################
# Elasticsearch queries #
#########################


def make_match_query(field: str, value: any) -> Dict[str, any]:
    """Make a basic ES match query."""
    return {
        'query': {
            'match': {
                field: value
            }
        }
    }


def make_scan_page_query(scan_id: str) -> Dict[str, any]:
    """Make a query to retrieve pages that have a given scan id."""
    return {
        'query': {
            'match': {
                'columns.metadata.scan_id.keyword': scan_id
            }
        }
    }


def get_random_query(size: int) -> Dict[str, any]:
    """Make a query for returning retrieval results in random order."""
    return {
        "query": {
            "function_score": {
                "random_score": {
                    "field": "_seq_no"
                }
            }
        },
        'size': size
    }


def retrieve_random_resolution_pages(es: Elasticsearch, config: Dict[str, any],
                                     size: int) -> List[ResolutionPageDoc]:
    """Retrieve a random set of resolution pages."""
    query = get_random_query(size)
    pages = rep_es.retrieve_pages_by_query(es, query, config)
    pages = [page for page in pages if page.metadata['page_type'] == 'resolution_page']
    return pages


def sample_scan_ids(es: Elasticsearch, config: Dict[str, any], size: int) -> List[str]:
    """Randomly sample scan ids by retrieving a random sample of indexed resolution pages."""
    pages = retrieve_random_resolution_pages(es, config, size)
    return [page.metadata['scan_id'] for page in pages]


#######################
# Spreadsheet parsing #
#######################


def initialise_workbook() -> Workbook:
    """Initialise an OpenPyXL Workbook with headers for resolution ground truth"""
    wb = Workbook()
    ws = wb.active
    headers = [
        'scan_id',
        'page_num',
        'line_id',
        'line_left',
        'line_right',
        'line_top',
        'line_bottom',
        'text',
        'res_start',
        'iiif_url',
        'notes'
    ]
    ws.append(headers)
    return wb


def pages_to_worksheet_rows(pages: List[ResolutionPageDoc],
                            res_start_ids: List[str]) -> Generator[List[str], None, None]:
    """Transform a list of resolution pages in a list of worksheet rows corresponding to
    page lines."""
    for page in pages:
        for line in page.stream_ordered_lines(include_header=True):
            res_start = 1 if line['metadata']['id'] in res_start_ids else ''
            iiif_url = f'=HYPERLINK("{page.metadata["iiif_url"]}", "iiif_url")'
            row = [
                page.metadata['scan_id'], page.metadata['page_num'], line['metadata']['id'],
                line['coords']['left'], line['coords']['right'],
                line['coords']['top'], line['coords']['bottom'],
                line['text'], res_start, iiif_url
            ]
            yield row
        # Add an empty row between pages for easy spotting of pages boundaries
        yield []
    return None


def scan_ids_to_grount_truth(es: Elasticsearch, scan_ids: List[str], config: Dict[str, any],
                             ground_truth_file: str) -> None:
    """Generate a ground truth excel sheet based on a list of resolution scan ids."""
    wb = initialise_workbook()
    ws = wb.active
    for scan_id in scan_ids:
        pages = get_scan_pages(es, scan_id, config)
        res_start_ids = get_scan_resolution_start_line_ids(es, scan_id, config)
        for row in pages_to_worksheet_rows(pages, res_start_ids):
            ws.append(row)
        # Add two empty rows between scans for easy spotting of scan boundaries
        ws.append([])
        ws.append([])
    wb.save(ground_truth_file)
    return None


def make_ground_truth_filename(ground_truth_dir: str, file_num: int) -> str:
    ground_truth_file = f'republic_ground_truth.res_start.{file_num}.xlsx'
    return os.path.join(ground_truth_dir, ground_truth_file)


def generate_ground_truth_file(es: Elasticsearch, config: Dict[str, any],
                               ground_truth_dir: str, size: int) -> None:
    file_num = get_next_file_num(ground_truth_dir)
    scan_ids = sample_scan_ids(es, config, size)
    ground_truth_file = make_ground_truth_filename(ground_truth_dir, file_num)
    scan_ids_to_grount_truth(es, scan_ids, config, ground_truth_file)


def get_next_file_num(ground_truth_dir: str) -> int:
    existing_files = glob.glob(ground_truth_dir + 'republic_ground_truth.res_start.*.xlsx')
    return len(existing_files) + 1


def get_scan_pages(es: Elasticsearch, scan_id: str, config: Dict[str, any]) -> List[ResolutionPageDoc]:
    query = make_match_query('metadata.scan_id.keyword', scan_id)
    return rep_es.retrieve_pages_by_query(es, query, config)


def get_scan_resolutions(es: Elasticsearch, scan_id: str,
                         config: Dict[str, any]) -> List[Resolution]:
    query = make_match_query('columns.metadata.scan_id.keyword', scan_id)
    return rep_es.retrieve_resolutions_by_query(es, query, config)


def get_resolution_first_line(resolution: Resolution) -> Dict[str, any]:
    col = resolution.columns[0]
    tr = col['textregions'][0]
    line = tr['lines'][0]
    return line


def get_scan_first_resolution_lines(es: Elasticsearch, scan_id: str,
                                    config: Dict[str, any]) -> List[Dict[str, any]]:
    resolutions = get_scan_resolutions(es, scan_id, config)
    first_lines = [get_resolution_first_line(resolution) for resolution in resolutions]
    return [first_line for first_line in first_lines if first_line['metadata']['scan_id'] == scan_id]


def get_scan_resolution_start_line_ids(es: Elasticsearch, scan_id: str,
                                       config: Dict[str, any]) -> List[str]:
    scan_first_lines = get_scan_first_resolution_lines(es, scan_id, config)
    return [line['metadata']['id'] for line in scan_first_lines]


def get_gt_sheet(gt_file: str):
    wb = load_workbook(filename=gt_file)
    return wb['Sheet']


def parse_gt_sheet(gt_sheet) -> List[Dict[str, any]]:
    headers = None
    lines = []
    for ri, row in enumerate(gt_sheet):
        if ri == 0:
            headers = [cell.value for cell in row]
            continue
        line_json = {headers[ci]: cell.value for ci, cell in enumerate(row)}
        lines.append(line_json)
    return lines


def parse_gt_file(gt_file: str):
    gt_sheet = get_gt_sheet(gt_file)
    lines = parse_gt_sheet(gt_sheet)
    return lines


def reformat_sheet(es: Elasticsearch, ground_truth_file: str, config: Dict[str, any]):
    lines = parse_gt_file(ground_truth_file)
    res_start_lines = [line for line in lines if line['res_start'] and line['res_start'] == 1]
    res_start_coords = [f"{line['scan_id']}-{line['line_left']}-{line['line_top']}" for line in res_start_lines]
    wb = initialise_workbook()
    ws = wb.active
    scan_ids: Set[str] = {line['scan_id'] for line in lines if line['scan_id']}
    for scan_id in scan_ids:
        pages = get_scan_pages(es, scan_id, config)
        res_start_ids = []
        for row in pages_to_worksheet_rows(pages, res_start_ids):
            if len(row) > 0:
                line_coord = f"{row[0]}-{row[3]}-{row[5]}"
                if line_coord in res_start_coords:
                    row[8] = 1
                    print(row)
            ws.append(row)
        # Add two empty rows between scans for easy spotting of scan boundaries
        ws.append([])
        ws.append([])
    # wb.save(ground_truth_file)

